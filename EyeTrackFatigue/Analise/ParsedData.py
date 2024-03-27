import math
from pandas import read_excel, concat, DataFrame
from openpyxl import load_workbook
import csv
import numpy as np
from ..Analise.Movements import EyeMovement
from ..Analise.GeoMetrics import GeoMetrics
from ..Analise.SecondaryMetrics import SecondaryMetrics

# основной класс, объединяет методы вычисления, обработки и хранение данных о количественных характеристиках
class ParsedData:
    def __init__(self):
        self.dots = 0
        self.raw_data = None
        self.secondary = None
        self.time_frame = 0
        self.minSpeed = 0
        self.avrSpeed = 0
        self.maxSpeed = 0
        self.fixations = []
        self.saccades = []
        self.flag = []
        self.pong = 0
        self.false_fix = 0
        self.method = 0
        self.args = 200
        self.center_radius = 200

    def compare(self, other): # метод сравнения двух отрезков обработанных данных 
        l = min(len(self.flag), len(other.flag))
        fp = 0
        fn = 0
        tp = 0
        for i in range(0, l):
            if self.flag[i] == 's' and other.flag[i] != 's':
                fp += 1.0
            if other.flag[i] == 's':
                if self.flag[i] == other.flag[i]:
                    tp += 1.0
                else:
                    fn += 1.0
        return 2 * tp / (2 * tp + fp + fn)

    def parse(self, input_section, method=0, args=200, center_radius=2.5): # разметка входных данных (выделение саккад и фиксаций) указанным методом
        print('parse...')
        # предварительная инициализация, выделение базовых характеристик
        self.raw_data = input_section
        self.center_radius = center_radius
        self.dots = len(input_section.positionData)
        self.sep_geo(input_section.FOV)
        self.time_frame = input_section.time_frame()
        self.minSpeed = min(input_section.velocityData)
        self.maxSpeed = max(input_section.velocityData)
        self.avrSpeed = sum(input_section.distanceData) / self.time_frame
        self.method = method
        self.args = args
        print('marking fixations...')
        self.flag = []
        # методы разметки данных
        if self.method != 2 and self.method != 1: 
            self.method = 0
        if self.method == 2:
            self.get_fixation_by_abs_dist(args)
        if self.method == 1:
            self.get_fixation_by_area(args)
        if self.method == 0:
            self.get_fixation_by_speed_threshold(args)
        print('cleaning...')
        self.clean_fix() # отчистка данных
        print('Pong = ' + self.pong.__str__() + ' False Fix = ' + self.false_fix.__str__())
        print('composing...')
        self.compose() # сбор размеченных данных в набор стратегий движения

    def parse_pupil(self, input_section, file): # разметка на основе внешних данных
        print('parse by pupil...')
        self.raw_data = input_section
        # предварительная инициализация, выделение базовых характеристик
        self.center_radius = 0
        self.dots = len(input_section.positionData)
        self.time_frame = input_section.time_frame()
        self.minSpeed = min(input_section.velocityData)
        self.maxSpeed = max(input_section.velocityData)
        self.avrSpeed = sum(input_section.distanceData) / self.time_frame
        self.method = 4
        self.args = file
        print('marking fixations...')
        self.flag = []
        self.get_fixation_by_pupil(file)
        print('cleaning...')
        self.clean_fix() # отчистка данных
        print('Pong = ' + self.pong.__str__() + ' False Fix = ' + self.false_fix.__str__())
        print('composing...')
        self.compose() # сбор размеченных данных в набор стратегий движения

    def get_method(self): # запрос выбранного метода разметки
        if self.method == 4:
            return 'by pupil_labs file '
        if self.method == 2:
            return 'abs dist ' + self.args.__str__()
        if self.method == 1:
            return 'area size ' + self.args.__str__()
        self.method = 0
        return 'speed ' + self.args.__str__()
    
    # Выделение нормированного X и Y 
    def sep_geo(self, FOV):  
        self.dots_X = []
        self.dots_Y = []
        for i in range(0, len(self.raw_data.positionData)):
            self.dots_X.append(self.raw_data.positionData[i].x / FOV)
            self.dots_Y.append(self.raw_data.positionData[i].y / FOV)

    # Алгоритм 1. Определение стратегий по скорости движения глаз между любыми двумя последовательными точками,
    # (Shic, Scassellati, & Chawarska, 2008).

    def get_fixation_by_speed_threshold(self, max_speed):
        print('max_speed ' + max_speed.__str__())
        self.flag.append('f')
        for i in range(0, self.dots - 1):
            if self.raw_data.velocityData[i] > max_speed:
                self.flag.append('s')
            else:
                self.flag.append('f')
    
    # Алгоритм 2. Определение стратегий по расстоянию между точками и центром фиксации, т.е. её радиусу.
    # (Camilli, Terenzi, & Nocera, 2008)

    def get_fixation_by_area(self, area_size):
        fix_area = self.raw_data.positionData[0]
        self.flag.append('f')
        for i in range(1, self.dots - 1):
            if self.flag[-1] == 'f':
                if fix_area.get_distance(self.raw_data.positionData[i]) > (area_size / 2):
                    self.flag.append('s')
                else:
                    self.flag.append('f')
            else:
                if self.raw_data.positionData[i - 1].get_distance(self.raw_data.positionData[i]) > (area_size / 2):
                    self.flag.append('s')
                else:
                    self.flag.append('f')
                    fix_area = self.raw_data.positionData[i - 1]

    # Алгоритм 3. Определение стратегий по максимальному расстоянию по горизонтали и вертикали,
    # покрываемому позициями взгляда при фиксации
    # Algorithm 3 Determination of strategies based on the maximum horizontal and vertical distance,
    # covered by the gaze positions inside fixation
    # (Salvucci & Goldberg). , 2000).

    def get_fixation_by_abs_dist(self, dist):
        max_x = min_x = self.raw_data.positionData[0].x
        max_y = min_y = self.raw_data.positionData[0].y
        self.flag.append('f')
        for i in range(1, self.dots - 1):
            max_x = max(max_x, self.raw_data.positionData[i].x)
            min_x = min(min_x, self.raw_data.positionData[i].x)
            max_y = max(max_y, self.raw_data.positionData[i].y)
            min_y = min(min_y, self.raw_data.positionData[i].y)
            if math.sqrt(pow(max_x - min_x, 2) + pow(max_y - min_y, 2)) > dist: # проверка порога / treshhold check
                self.flag.append('s')
                max_x = min_x = self.raw_data.positionData[i].x
                max_y = min_y = self.raw_data.positionData[i].y
            else:
                self.flag.append('f')

    # разметка по внешним данным (например используя файл полученный от обработки данных алгоритмами PupilLabs)
    def get_fixation_by_pupil(self, file):
        fix = []
        with open(file, newline='') as f:
            reader = csv.reader(f)
            head = True
            start_time = 0
            for row in reader:
                if head:
                    head = False
                    continue
                if start_time == 0:
                    start_time = float(row[3]) / 1000000000
                fix.append([float(row[3]) / 1000000000 - start_time, float(row[4]) / 1000000000 - start_time])

        print("Pupil time" + (fix[-1][1] - fix[0][0]).__str__())
        print("Frame time" + self.time_frame.__str__())
        delta_time = abs(self.time_frame - (fix[-1][1] - fix[0][0]))
        #start_time -= delta_time
        fix_i = 0
        while self.raw_data.positionData[0].time > fix[fix_i + 1][0]:
            fix_i += 1
        print(fix_i.__str__() + '  ' + self.raw_data.positionData[0].time.__str__() + '  '
              + fix[fix_i][0].__str__(), fix[fix_i][1].__str__())
        # print(fix)
        i = 0
        while i < self.dots:
            if fix_i >= len(fix):
                print('sss')
                break
            if self.raw_data.positionData[i].time < fix[fix_i][0]:
                self.flag.append('s')
            elif self.raw_data.positionData[i].time < fix[fix_i][1]:
                self.flag.append('f')
            else:
                fix_i += 1
                continue
            i += 1

    def clean_fix(self): # выделение "ложных" (сверхкоротких) фиксаций и саккад
        for i in range(1, len(self.flag) - 2):
            if self.flag[i] == 's' and self.flag[i - 1] == 'f' and self.flag[i + 1] == 'f':
                self.pong += 1
                self.flag[i] = 'f'
        for i in range(1, len(self.flag) - 2):
            if self.flag[i] == 'f' and self.flag[i - 1] == 's' and self.flag[i + 1] == 's':
                self.false_fix += 1
                self.flag[i] = 's'
    
    def compose(self):  # сбор размеченных данных в набор стратегий движения
        fl = self.flag[0]
        self.new_movement(fl, self.raw_data.positionData[0])
        for i in range(1, len(self.flag) - 1):
            if fl == 's':
                self.saccades[-1].add_point(self.raw_data.positionData[i])
            else:
                if fl == self.flag[i]:
                    self.fixations[-1].add_point(self.raw_data.positionData[i])
            if fl != self.flag[i]:
                fl = self.flag[i]
                if fl == 's':
                    self.new_movement(fl, self.raw_data.positionData[i-1])
                    self.saccades[-1].add_point(self.raw_data.positionData[i])
                else:
                    self.new_movement(fl, self.raw_data.positionData[i])

    def new_movement(self, fl, point): # добавление новой стратегии движения
        if fl == 's':
            self.saccades.append(EyeMovement('s'))
            self.saccades[-1].add_point(point)
        else:
            self.fixations.append(EyeMovement('f'))
            self.fixations[-1].add_point(point)

    def calc_metrics(self): # рассчёт характеристик
        # предварительная инициализация
        self.secondary = SecondaryMetrics() 
        self.geo = GeoMetrics()

        # рассчёт геометрических метрик позиции по оси Х
        x = np.array(self.dots_X)
        self.geo.x_mean = np.mean(x).reshape(1, 1)[0][0]
        self.geo.x_std = np.std(x).reshape(1, 1)[0][0]
        self.geo.x_min = np.min(x).reshape(1, 1)[0][0]
        self.geo.x_max = np.max(x).reshape(1, 1)[0][0]
        self.geo.x_25 = np.percentile(x, 25).reshape(1, 1)[0][0]
        self.geo.x_50 = np.percentile(x, 50).reshape(1, 1)[0][0]
        self.geo.x_75 = np.percentile(x, 75).reshape(1, 1)[0][0]
        # рассчёт геометрических метрик позиции по оси Y
        y = np.array(self.dots_Y)
        self.geo.y_mean = np.mean(y).reshape(1, 1)[0][0]
        self.geo.y_std = np.std(y).reshape(1, 1)[0][0]
        self.geo.y_min = np.min(y).reshape(1, 1)[0][0]
        self.geo.y_max = np.max(y).reshape(1, 1)[0][0]
        self.geo.y_25 = np.percentile(y, 25).reshape(1, 1)[0][0]
        self.geo.y_50 = np.percentile(y, 50).reshape(1, 1)[0][0]
        self.geo.y_75 = np.percentile(y, 75).reshape(1, 1)[0][0]

        print('computing metrics...')
        print('fixations...')
        
        for fix in self.fixations: # рассчёт характеристик для фиксаций
            if len(fix.points) < 3:
                continue
            self.secondary.fix_time += fix.time
            self.secondary.fix_distance += fix.get_distance()
            self.secondary.max_f_speed = max(self.secondary.max_f_speed, fix.get_avr_speed())
            self.secondary.min_f_speed = min(self.secondary.min_f_speed, fix.get_avr_speed())
            # self.secondary.min_curve = min(self.secondary.min_curve, fix.get_curve())
            # self.secondary.max_curve = min(self.secondary.max_curve, fix.get_curve())
            # self.secondary.avr_curve += fix.get_curve()
            if fix.time > 0.180: # выделение данных о фиксациях дольше 180мс
                self.secondary.fix_g_180 += 1
                self.secondary.fix_g_180_time += fix.time
                self.secondary.fix_g_180_count += 1
            if fix.time < 0.180: # выделение данных о фиксациях короче 180мс
                self.secondary.fix_l_180 += 1
                self.secondary.fix_l_180_time += fix.time
                self.secondary.fix_l_180_count += 1 
            if fix.time < 0.150: # выделение данных о фиксациях короче 150мс
                self.secondary.short_fix_time += fix.time
                self.secondary.short_fix_count += 1
                self.secondary.fix_l_150_dist += fix.get_distance()
                self.secondary.max_sf_speed = max(self.secondary.max_sf_speed, fix.get_avr_speed())
                self.secondary.min_sf_speed = min(self.secondary.min_sf_speed, fix.get_avr_speed())
            if 0.150 <= fix.time <= 0.900: # выделение данных о фиксациях дольше 150мс но короче 900мс
                self.secondary.med_fix_time += fix.time
                self.secondary.med_fix_count += 1
                self.secondary.fix_g_150_dist += fix.get_distance()
                self.secondary.max_lf_speed = max(self.secondary.max_lf_speed, fix.get_avr_speed())
                self.secondary.min_lf_speed = min(self.secondary.min_lf_speed, fix.get_avr_speed())
            if fix.time > 0.900: # выделение данных о фиксациях дольше 900мс
                self.secondary.long_fix_time += fix.time
                self.secondary.long_fix_count += 1
                self.secondary.fix_g_150_dist += fix.get_distance()
                self.secondary.max_lf_speed = max(self.secondary.max_lf_speed, fix.get_avr_speed())
                self.secondary.min_lf_speed = min(self.secondary.min_lf_speed, fix.get_avr_speed())

        print('saccades...')
        for sacc in self.saccades: # рассчёт характеристик для саккад
            if len(sacc.points) < 3:
                continue
            self.secondary.sacc_time += sacc.time
            self.secondary.sacc_distance += sacc.get_distance()
            self.secondary.max_s_speed = max(self.secondary.max_s_speed, sacc.get_avr_speed())
            self.secondary.min_s_speed = min(self.secondary.min_s_speed, sacc.get_avr_speed())
            self.secondary.max_curve = max(self.secondary.max_curve, sacc.get_curve())
            self.secondary.min_curve = min(self.secondary.min_curve, sacc.get_curve())
            self.secondary.avr_curve += sacc.get_curve()
            self.secondary.max_s_length = max(self.secondary.max_s_length, sacc.get_distance())
            self.secondary.min_s_length = min(self.secondary.min_s_length, sacc.get_distance())
            self.secondary.max_s_time = max(self.secondary.max_s_time, sacc.time)
            self.secondary.min_s_time = min(self.secondary.min_s_time, sacc.time)
            self.secondary.sacc_length += sacc.get_distance()
            if sacc.get_abs_distance() > 6:
                self.secondary.long_sacc += 1
            else:
                self.secondary.short_sacc += 1


        # инициализация переменных для рассчёта дополнительных характеристик
        i_count = 0
        s0 = self.raw_data.positionData[0]
        t0 = s0.time
        v0 = 0
        prev = s0
        dist = 0
        v = -1
        p = None
        # рассчёт дополнительных характеристик
        for step in self.raw_data.positionData:
            if p is not None:
                if v != -1:
                    acc = abs(v - p.get_distance(step) / (step.time - p.time))  # / (step.time - p.time)
                    self.secondary.avr_acc += acc
                    self.secondary.min_acc = min(self.secondary.min_i_acc, acc / (step.time - p.time))
                    self.secondary.max_acc = max(self.secondary.max_i_acc, acc / (step.time - p.time))
                v = p.get_distance(step) / (step.time - p.time)
            p = step

            if step.time - t0 > 1:
                dt = step.time - t0
                dv = abs(v0 - dist/dt)
                self.secondary.avr_i_speed += dist / dt
                self.secondary.min_i_speed = min(self.secondary.min_i_speed, dist / dt)
                self.secondary.max_i_speed = max(self.secondary.max_i_speed, dist / dt)
                self.secondary.avr_i_acc += dv / dt
                self.secondary.min_i_acc = min(self.secondary.min_i_acc, dv / dt)
                self.secondary.max_i_acc = max(self.secondary.max_i_acc, dv / dt)
                i_count += 1
                s0 = step
                t0 = s0.time
                v0 = dist / dt
                dist = 0
            dist += step.get_distance(prev)
            prev = step


        self.secondary.avr_acc /= self.raw_data.positionData[-1].time - self.raw_data.positionData[2].time
        self.secondary.avr_i_speed /= i_count
        self.secondary.avr_i_acc /= i_count

        i_count = 0
        t0 = self.raw_data.positionData[0].time
        fix_count = 0
        for i in range(1, len(self.flag)):
            if self.flag[i] == 'f' and self.flag[i-1] == 's':
                fix_count += 1
            if self.raw_data.positionData[i].time - t0 > 1:
                freq = fix_count / (self.raw_data.positionData[i].time - t0)
                self.secondary.avr_freq += freq
                self.secondary.min_freq = min(self.secondary.min_freq, freq)
                self.secondary.max_freq = max(self.secondary.max_freq, freq)
                i_count += 1
                t0 = self.raw_data.positionData[i].time
                fix_count = 0
        self.secondary.avr_freq /= i_count
        print('calculating complete')

    def get_row(self, file_name):   # запрос характеристик в виде списка 
        print('forming a row... ' )
        return self.get_df_row(file_name).loc[0, :].values.flatten().tolist()

    def to_csv(self, file_name, file_to): # вывод данных в файл формата .csv
        print('csv file... ' + file_to)
        row = self.get_df_row(file_name)
        if os.path.exists(file_to):
            row.to_csv(file_to, mode='a', index=False, header=False, sep=';')
        else:
            row.to_csv(file_to, index=False, header=True, sep=';')

    def get_df_row(self, file='none'): # запрос характеристик в виде DataFrame
        all_fix = self.secondary.short_fix_count + self.secondary.med_fix_count + self.secondary.long_fix_count
        df_row = DataFrame({'File' : [file],
                            'x_mean' : [self.geo.x_mean],
                            'x_std' : [self.geo.x_std],
                            'x_min' : [self.geo.x_min],
                            'x_max' : [self.geo.x_max],
                            'x_25' : [self.geo.x_25],
                            'x_50' : [self.geo.x_50],
                            'x_75' : [self.geo.x_75],
                            'y_mean' : [self.geo.y_mean],
                            'y_std' : [self.geo.y_std],
                            'y_min' : [self.geo.y_min],
                            'y_max' : [self.geo.y_max],
                            'y_25' : [self.geo.y_25],
                            'y_50' : [self.geo.y_50],
                            'y_75' : [self.geo.y_75],
                            'False Fixation, per minute': [self.false_fix / self.time_frame * 60],
                            'False Saccades, per minute': [self.pong / self.time_frame * 60],
                            'Saccades with amplitude > 6 degrees, per minute': [
                                self.secondary.long_sacc / self.time_frame * 60],
                            'Saccades with amplitude < 6 degrees, per minute': [
                                self.secondary.short_sacc / self.time_frame * 60],
                            'Max Curve': [self.secondary.max_curve],
                            'Average Curve': [self.secondary.avr_curve / (len(self.fixations) + len(self.saccades))],
                            'Min Curve': [self.secondary.min_curve],
                            'Time Frame': [self.time_frame],
                            'Saccades time': [self.secondary.sacc_time],
                            'Fixation time < 150 ms': [self.secondary.short_fix_time],
                            'Fixation time > 150 ms': [self.secondary.med_fix_time + self.secondary.long_fix_time],
                            'Fixation time between 150 and 900 ms': [self.secondary.med_fix_time],
                            'Fixation time > 900 ms': [self.secondary.long_fix_time],
                            'Fixation time < 180 ms': [self.secondary.fix_l_180_time],
                            'Fixation time > 180 ms': [self.secondary.fix_g_180_time],
                            '% of Fixations < 150 ms': [self.secondary.short_fix_count / all_fix * 100 if all_fix > 0 else -1],
                            '% of Fixations > 150 ms': [(1 - self.secondary.short_fix_count / all_fix) * 100 if all_fix > 0 else -1],
                            '% of Fixations between 150 and 900 ms': [self.secondary.med_fix_count / all_fix * 100 if all_fix > 0 else -1],
                            '% of Fixations > 900 ms': [self.secondary.long_fix_count / all_fix * 100 if all_fix > 0 else -1],
                            '% of Fixations < 180 ms': [self.secondary.fix_l_180 / all_fix * 100 if all_fix > 0 else -1],
                            '% of Fixations > 180 ms': [self.secondary.fix_g_180 / all_fix * 100 if all_fix > 0 else -1],
                            'Fixation time < 150 ms, per time': [self.secondary.short_fix_time / self.time_frame],
                            'Fixation time > 150 ms, per time': [
                                (self.secondary.med_fix_time + self.secondary.long_fix_time) / self.time_frame],
                            'Fixation time between 150 and 900 ms, per time': [
                                self.secondary.med_fix_time / self.time_frame],
                            'Fixation time > 900 ms, per time': [self.secondary.long_fix_time / self.time_frame],
                            'Fixation time < 180 ms, per time': [self.secondary.fix_l_180_time / self.time_frame],
                            'Fixation time > 180 ms, per time': [self.secondary.fix_g_180_time / self.time_frame],
                            '% of Fixations < 150 ms, per minute': [
                                self.secondary.short_fix_count / self.time_frame * 60],
                            '% of Fixations > 150 ms, per minute': [
                                (self.secondary.med_fix_count + self.secondary.long_fix_count) / self.time_frame * 60],
                            '% of Fixations between 150 and 900 ms, per minute': [
                                self.secondary.med_fix_count / self.time_frame * 60],
                            '% of Fixations > 900 ms, per minute': [
                                self.secondary.long_fix_count / self.time_frame * 60],
                            '% of Fixations < 180 ms, per minute': [self.secondary.fix_l_180 / self.time_frame * 60],
                            '% of Fixations > 180 ms, per minute': [self.secondary.fix_g_180 / self.time_frame * 60],
                            'Fixation Frequency': [len(self.fixations) / self.time_frame],
                            'Average Fix Frequency in interval (1s)': [self.secondary.avr_freq],
                            'Max Fix Frequency in interval (1s)': [self.secondary.max_freq],
                            'Average Acceleration': [self.secondary.avr_acc],
                            'Min Acceleration': [self.secondary.min_acc],
                            'Max Acceleration': [self.secondary.max_acc],
                            'Average Speed': [self.avrSpeed],
                            'Min Speed': [self.minSpeed],
                            'Max Speed': [self.maxSpeed],
                            'Average Acceleration in interval (1s)': [self.secondary.avr_i_acc],
                            'Min Acceleration in interval (1s)': [self.secondary.min_i_acc],
                            'Max Acceleration in interval (1s)': [self.secondary.max_i_acc],
                            'Average Speed in interval (1s)': [self.secondary.avr_i_speed],
                            'Min Speed in interval (1s)': [self.secondary.min_i_speed],
                            'Max Speed in interval (1s)': [self.secondary.max_i_speed],
                            'Average Fixation Speed': [self.secondary.fix_distance / self.secondary.fix_time if self.secondary.fix_time > 0 else - 1],
                            'Min Fixation Speed': [self.secondary.min_f_speed],
                            'Max Fixation Speed': [self.secondary.max_f_speed],
                            'Average Fixation Speed, < 150ms': [self.secondary.fix_l_150_dist / self.secondary.short_fix_time if self.secondary.short_fix_time > 0 else - 1],
                            'Min Fixation Speed, < 150ms': [self.secondary.min_sf_speed],
                            'Max Fixation Speed, < 150ms': [self.secondary.max_sf_speed],
                            'Average Fixation Speed, > 150ms': [self.secondary.fix_g_150_dist / (self.secondary.med_fix_time + self.secondary.long_fix_time) if (self.secondary.med_fix_time + self.secondary.long_fix_time) != 0 else -1],
                            'Min Fixation Speed, > 150ms': [self.secondary.min_lf_speed],
                            'Max Fixation Speed, > 150ms': [self.secondary.max_lf_speed],
                            'Average Saccade Speed': [self.secondary.sacc_distance / self.secondary.sacc_time if self.secondary.sacc_time > 0 else - 1],
                            'Min Saccade Speed': [self.secondary.min_s_speed],
                            'Max Saccade Speed': [self.secondary.max_s_speed],
                            'Average Saccade Length': [self.secondary.sacc_distance / len(self.saccades) if len(self.saccades) > 0 else - 1],
                            'Min Saccade Length': [self.secondary.min_s_length],
                            'Max Saccade Length': [self.secondary.max_s_length],
                            'Average Saccade Time': [self.secondary.sacc_time / len(self.saccades) if len(self.saccades) > 0 else - 1],
                            'Min Saccade Time': [self.secondary.min_s_time],
                            'Max Saccade Time': [self.secondary.max_s_time]
                            })
        return df_row

    def to_xls_by_row(self, file_name, file_to): # вывод данных в файл формата excel файлов, по рядам (строкам)
        print('excel reading file ' + file_to)
        print('excel file preparing ' + file_to)
        print('forming headline...')

        df = read_excel(file_to, engine='openpyxl')
        print(df.__str__())
        print('computing metrics...')
        print('forming a row...')
        df_row = self.get_df_row(file_name)

        print('appending a row...')
        df = concat([df, df_row], ignore_index=True)
        print(df.__str__())
        print('writing...')
        update_spreadsheet(file_to, df, df_row, sheet_name='Sheet1')
        print('writen')

    def to_xls(self, file_name, file_to): # вывод данных в файл формата excel файлов, путём обноления ячеек
        # может использоваться при необходимости частичного исправления данных в уже имеющихся таблицах
        # однако крайне не эффективен по временным затратам
        print('Not effective')
        print('excel reading file ' + file_to)
        print('excel file preparing ' + file_to)
        print('forming headline...')

        df = read_excel(file_to, engine='openpyxl')
        print(df.__str__())
        all_fix = self.secondary.short_fix_count + self.secondary.med_fix_count + self.secondary.long_fix_count
        new_row = len(df.get('File')) + 1
        update_cell(file_to, new_row, key_index(df, 'File'), file_name)
        update_cell(file_to, new_row, key_index(df, 'Method'), self.get_method())

        update_cell(file_to, new_row, key_index(df, 'False Fixation, per minute'),
                    self.false_fix / self.time_frame * 60)
        update_cell(file_to, new_row, key_index(df, 'False Saccades, per minute'),
                    self.pong / self.time_frame * 60)

        update_cell(file_to, new_row, key_index(df, 'Saccades with amplitude > 6 degrees, per minute'),
                    self.secondary.long_sacc / self.time_frame * 60)
        update_cell(file_to, new_row, key_index(df, 'Saccades with amplitude < 6 degrees, per minute'),
                    self.secondary.short_sacc / self.time_frame * 60)

        update_cell(file_to, new_row, key_index(df, 'Max Curve'), self.secondary.max_curve)
        update_cell(file_to, new_row, key_index(df, 'Average Curve'),
                    self.secondary.avr_curve / (len(self.fixations) + len(self.saccades)))
        update_cell(file_to, new_row, key_index(df, 'Min Curve'), self.secondary.min_curve)

        update_cell(file_to, new_row, key_index(df, 'Time Frame'), self.time_frame)
        update_cell(file_to, new_row, key_index(df, 'Saccades time'), self.secondary.sacc_time)

        update_cell(file_to, new_row, key_index(df, 'Fixation time < 150 ms'),
                    self.secondary.short_fix_time)
        update_cell(file_to, new_row, key_index(df, 'Fixation time > 150 ms'),
                    self.secondary.med_fix_time + self.secondary.long_fix_time)

        update_cell(file_to, new_row, key_index(df, 'Fixation time between 150 and 900 ms'),
                    self.secondary.med_fix_time)
        update_cell(file_to, new_row, key_index(df, 'Fixation time > 900 ms'),
                    self.secondary.long_fix_time)

        update_cell(file_to, new_row, key_index(df, 'Fixation time < 180 ms'),
                    self.secondary.fix_l_180_time)
        update_cell(file_to, new_row, key_index(df, 'Fixation time > 180 ms'),
                    self.secondary.fix_g_180_time)

        update_cell(file_to, new_row, key_index(df, '% of Fixations < 150 ms'),
                    self.secondary.short_fix_count / all_fix * 100)
        update_cell(file_to, new_row, key_index(df, '% of Fixations > 150 ms'),
                    (1 - self.secondary.short_fix_count / all_fix) * 100)

        update_cell(file_to, new_row, key_index(df, '% of Fixations between 150 and 900 ms'),
                    self.secondary.med_fix_count / all_fix * 100)

        update_cell(file_to, new_row, key_index(df, '% of Fixations > 900 ms'),
                    self.secondary.long_fix_count / all_fix * 100)

        update_cell(file_to, new_row, key_index(df, '% of Fixations < 180 ms'),
                    self.secondary.fix_l_180 / all_fix * 100)

        update_cell(file_to, new_row, key_index(df, '% of Fixations > 180 ms'),
                    self.secondary.fix_g_180 / all_fix * 100)

        # , per minute

        update_cell(file_to, new_row, key_index(df, 'Fixation time < 150 ms, per time'),
                    self.secondary.short_fix_time / self.time_frame)
        update_cell(file_to, new_row, key_index(df, 'Fixation time > 150 ms, per time'),
                    (self.secondary.med_fix_time + self.secondary.long_fix_time) / self.time_frame)
        update_cell(file_to, new_row, key_index(df, 'Fixation time between 150 and 900 ms, per time'),
                    self.secondary.med_fix_time / self.time_frame)
        update_cell(file_to, new_row, key_index(df, 'Fixation time > 900 ms, per time'),
                    self.secondary.long_fix_time / self.time_frame)

        update_cell(file_to, new_row, key_index(df, 'Fixation time < 180 ms, per time'),
                    self.secondary.fix_l_180_time / self.time_frame)
        update_cell(file_to, new_row, key_index(df, 'Fixation time > 180 ms, per time'),
                    self.secondary.fix_g_180_time / self.time_frame)

        update_cell(file_to, new_row, key_index(df, 'Fixations < 150 ms, per minute'),
                    self.secondary.short_fix_count / self.time_frame * 60)
        update_cell(file_to, new_row, key_index(df, 'Fixations > 150 ms, per minute'),
                    (self.secondary.med_fix_count + self.secondary.long_fix_count) / self.time_frame * 60)
        update_cell(file_to, new_row, key_index(df, 'Fixations between 150 and 900 ms, per minute'),
                    self.secondary.med_fix_count / self.time_frame * 60)
        update_cell(file_to, new_row, key_index(df, 'Fixations > 900 ms, per minute'),
                    self.secondary.long_fix_count / self.time_frame * 60)

        update_cell(file_to, new_row, key_index(df, 'Fixations < 180 ms, per minute'),
                    self.secondary.fix_l_180 / self.time_frame * 60)
        update_cell(file_to, new_row, key_index(df, 'Fixations > 180 ms, per minute'),
                    self.secondary.fix_g_180 / self.time_frame * 60)

        update_cell(file_to, new_row, key_index(df, 'Fixation Frequency'), len(self.fixations) / self.time_frame)
        update_cell(file_to, new_row, key_index(df, 'Average Fix Frequency in interval (1s)'), self.secondary.avr_freq)
        update_cell(file_to, new_row, key_index(df, 'Max Fix Frequency in interval (1s)'), self.secondary.max_freq)

        update_cell(file_to, new_row, key_index(df, 'Average Acceleration'), self.secondary.avr_acc)
        update_cell(file_to, new_row, key_index(df, 'Min Acceleration'), self.secondary.min_acc)
        update_cell(file_to, new_row, key_index(df, 'Max Acceleration'), self.secondary.max_acc)

        update_cell(file_to, new_row, key_index(df, 'Average Speed'), self.avrSpeed)
        update_cell(file_to, new_row, key_index(df, 'Min Speed'), self.minSpeed)
        update_cell(file_to, new_row, key_index(df, 'Max Speed'), self.maxSpeed)

        update_cell(file_to, new_row, key_index(df, 'Average Acceleration in interval (1s)'), self.secondary.avr_i_acc)
        update_cell(file_to, new_row, key_index(df, 'Min Acceleration in interval (1s)'), self.secondary.min_i_acc)
        update_cell(file_to, new_row, key_index(df, 'Max Acceleration in interval (1s)'), self.secondary.max_i_acc)

        update_cell(file_to, new_row, key_index(df, 'Average Speed in interval (1s)'), self.secondary.avr_i_speed)
        update_cell(file_to, new_row, key_index(df, 'Min Speed in interval (1s)'), self.secondary.min_i_speed)
        update_cell(file_to, new_row, key_index(df, 'Max Speed in interval (1s)'), self.secondary.max_i_speed)

        update_cell(file_to, new_row, key_index(df, 'Average Fixation Speed'),
                    self.secondary.fix_distance / self.secondary.fix_time)
        update_cell(file_to, new_row, key_index(df, 'Min Fixation Speed'), self.secondary.min_f_speed)
        update_cell(file_to, new_row, key_index(df, 'Max Fixation Speed'), self.secondary.max_f_speed)
        
        update_cell(file_to, new_row, key_index(df, 'Average Fixation Speed, < 150ms'),
                    self.secondary.fix_l_150_dist / self.secondary.short_fix_time)
        update_cell(file_to, new_row, key_index(df, 'Min Fixation Speed, < 150ms'), self.secondary.min_sf_speed)
        update_cell(file_to, new_row, key_index(df, 'Max Fixation Speed, < 150ms'), self.secondary.max_sf_speed)

        update_cell(file_to, new_row, key_index(df, 'Average Fixation Speed, > 150ms'),
                    self.secondary.fix_g_150_dist / (self.secondary.med_fix_time + self.secondary.long_fix_time))
        update_cell(file_to, new_row, key_index(df, 'Min Fixation Speed, > 150ms'), self.secondary.min_lf_speed)
        update_cell(file_to, new_row, key_index(df, 'Max Fixation Speed, > 150ms'), self.secondary.max_lf_speed)

        update_cell(file_to, new_row, key_index(df, 'Average Saccade Speed'),
                    self.secondary.sacc_distance / self.secondary.sacc_time)
        update_cell(file_to, new_row, key_index(df, 'Min Saccade Speed'), self.secondary.min_s_speed)
        update_cell(file_to, new_row, key_index(df, 'Max Saccade Speed'), self.secondary.max_s_speed)

        update_cell(file_to, new_row, key_index(df, 'Average Saccade Length'),
                    self.secondary.sacc_distance / len(self.saccades))
        update_cell(file_to, new_row, key_index(df, 'Min Saccade Length'), self.secondary.min_s_length)
        update_cell(file_to, new_row, key_index(df, 'Max Saccade Length'), self.secondary.max_s_length)

        update_cell(file_to, new_row, key_index(df, 'Average Saccade Time'),
                    self.secondary.sacc_time / len(self.saccades))
        update_cell(file_to, new_row, key_index(df, 'Min Saccade Time'), self.secondary.min_s_time)
        update_cell(file_to, new_row, key_index(df, 'Max Saccade Time'), self.secondary.max_s_time)

        update_cell(file_to, new_row, key_index(df, 'Average PPI'), 'none')
        update_cell(file_to, new_row, key_index(df, 'Min PPI'), 'none')
        update_cell(file_to, new_row, key_index(df, 'Max PPI'), 'none')

        print('writen')

    @staticmethod
    def ppi_to_xls(file_from, lines, file_to): # метод для добавления доп. информации о преимпульсном ингибировании
        # использование этих данных не показало ожидаемого результата, однако оставлена возможность для доработки / добавления другой дополнительной информации
        print('excel file preparing ' + file_to)
        print('forming headline...')
        df = read_excel(file_to, engine='openpyxl')
        print('finding a row...')
        i = -1
        temp = 1
        file_from = file_from.split('/')[-1]
        for file_name in df.get('File'):
            if file_name.split('.')[0] == file_from.split('.')[0]:
                i = temp
            temp += 1
        if i == -1:
            print('no such file')
            return

        print('forming ppi')
        ppi = []
        first_fl = True
        for line in lines:
            if first_fl:
                first_fl = False
                continue
            ppi.append(float(line.split(';')[1].split('\n')[0]))

        update_cell(file_to, i, key_index(df, 'Average PPI'), sum(ppi)/len(ppi))
        update_cell(file_to, i, key_index(df, 'Min PPI'), min(ppi))
        update_cell(file_to, i, key_index(df, 'Max PPI'), max(ppi))
        print('PPI writen')

    def __str__(self): # вывод в стоковом формате, выводитьcя частичная информация о размеченном отрезке
        s = '__Metrics:__\n'
        s += 'Time Frame: ' + self.time_frame.__str__() + ' Dots count: ' + self.dots.__str__() + '\n'
        s += 'Estimated Hrz: ' + (self.dots / self.time_frame).__str__() + '\n'
        s += 'False Fix: ' + self.false_fix.__str__() + '\t'
        s += 'Ping-pong: ' + self.pong.__str__() + '\n'
        s += 'Min Speed: ' + f'{self.minSpeed:.6f}' + '\n'
        s += 'Max Speed: ' + f'{self.maxSpeed:.6f}' + '\n'
        s += 'Avr Speed: ' + f'{self.avrSpeed:.6f}' + '\n'

        fix_distance = 0
        fix_time = 0
        min_fix_speed = 1000000
        max_fix_speed = 0

        fix_min_acc = 100000
        fix_max_acc = 0
        fix_avr_acc = 0

        for fix in self.fixations:
            fix_time += fix.time
            fix_distance += fix.get_abs_distance()
            min_fix_speed = min(min_fix_speed, fix.get_min_speed())
            max_fix_speed = max(max_fix_speed, fix.get_max_speed())
            fix_min_acc = min(fix_min_acc, fix.get_min_acceleration())
            fix_max_acc = max(fix_max_acc, fix.get_max_acceleration())
            fix_avr_acc += fix.get_avr_acceleration() * fix.time

        s += 'Fix Min Speed: ' + f'{min_fix_speed:.6f}' + '\n'
        s += 'Fix Max Speed: ' + f'{max_fix_speed:.6f}' + '\n'
        s += 'Fix Avr Speed: ' + f'{fix_distance / fix_time:.6f}' + '\n'
        s += 'Fix Min Acceleration: ' + f'{fix_min_acc:.6f}' + '\n'
        s += 'Fix Max Acceleration: ' + f'{fix_max_acc:.6f}' + '\n'
        s += 'Fix Avr Acceleration: ' + f'{fix_avr_acc / fix_time:.6f}' + '\n'
        return s

    def get_section(self):
        return self.raw_data

    @staticmethod
    def get_df_null_row(): # статический метод для запроса шаблона таблицы с рассчитываемыми характеристиками
        return DataFrame({'File' : [],
                          'x_mean' : [],
                          'x_std' : [],
                          'x_min' : [],
                          'x_max' : [],
                          'x_25' : [],
                          'x_50' : [],
                          'x_75' : [],
                          'y_mean' : [],
                          'y_std' : [],
                          'y_min' : [],
                          'y_max' : [],
                          'y_25' : [],
                          'y_50' : [],
                          'y_75' : [],
                          'False Fixation, per minute': [],
                          'False Saccades, per minute': [],
                          'Saccades with amplitude > 6 degrees, per minute': [],
                          'Saccades with amplitude < 6 degrees, per minute': [],
                          'Max Curve': [],
                          'Average Curve': [],
                          'Min Curve': [], 
                          'Time Frame': [],
                          'Saccades time': [],
                          'Fixation time < 150 ms': [],
                          'Fixation time > 150 ms': [],
                          'Fixation time between 150 and 900 ms': [],
                          'Fixation time > 900 ms': [],
                          'Fixation time < 180 ms': [],
                          'Fixation time > 180 ms': [],
                          '% of Fixations < 150 ms': [],
                          '% of Fixations > 150 ms': [],
                          '% of Fixations between 150 and 900 ms': [],
                          '% of Fixations > 900 ms': [],
                          '% of Fixations < 180 ms': [],
                          '% of Fixations > 180 ms': [],
                          'Fixation time < 150 ms, per time': [],
                          'Fixation time > 150 ms, per time': [],
                          'Fixation time between 150 and 900 ms, per time': [],
                          'Fixation time > 900 ms, per time': [],
                          'Fixation time < 180 ms, per time': [],
                          'Fixation time > 180 ms, per time': [],
                          '% of Fixations < 150 ms, per minute': [],
                          '% of Fixations > 150 ms, per minute': [],
                          '% of Fixations between 150 and 900 ms, per minute': [],
                          '% of Fixations > 900 ms, per minute': [],
                          '% of Fixations < 180 ms, per minute': [],
                          '% of Fixations > 180 ms, per minute': [],
                          'Fixation Frequency': [],
                          'Average Fix Frequency in interval (1s)': [],
                          'Max Fix Frequency in interval (1s)': [],
                          'Average Acceleration': [],
                          'Min Acceleration': [],
                          'Max Acceleration': [],
                          'Average Speed': [],
                          'Min Speed': [],
                          'Max Speed': [],
                          'Average Acceleration in interval (1s)': [],
                          'Min Acceleration in interval (1s)': [],
                          'Max Acceleration in interval (1s)': [],
                          'Average Speed in interval (1s)': [],
                          'Min Speed in interval (1s)': [],
                          'Max Speed in interval (1s)': [],
                          'Average Fixation Speed': [],
                          'Min Fixation Speed': [],
                          'Max Fixation Speed': [],
                          'Average Fixation Speed, < 150ms': [],
                          'Min Fixation Speed, < 150ms': [],
                          'Max Fixation Speed, < 150ms': [],
                          'Average Fixation Speed, > 150ms': [],
                          'Min Fixation Speed, > 150ms': [],
                          'Max Fixation Speed, > 150ms': [],
                          'Average Saccade Speed': [],
                          'Min Saccade Speed': [],
                          'Max Saccade Speed': [],
                          'Average Saccade Length': [],
                          'Min Saccade Length': [],
                          'Max Saccade Length': [],
                          'Average Saccade Time': [],
                          'Min Saccade Time': [],
                          'Max Saccade Time': []})

# метод для обновления excel таблиц
def update_spreadsheet(path: str, _df, _new_row, sheet_name: str = "Sheet1"):
    '''
    :param path: Путь до файла Excel
    :param _df: Датафрейм Pandas для записи
    :param starcol: Стартовая колонка в таблице листа Excel, куда буду писать данные
    :param startrow: Стартовая строка в таблице листа Excel, куда буду писать данные
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :return:
    '''
    wb = load_workbook(path)
    starcol = 1
    startrow = len(_df) + 1
# print(len(wb[sheet_name]))
    for ir in range(0, len(_new_row)):
        for ic in range(0, len(_new_row.iloc[ir])):
            wb[sheet_name].cell(startrow + ir, starcol + ic).value = _new_row.iloc[ir][ic]
    wb.save(path)

# метод для обновления конкретной ячейки excel таблицы
def update_cell(path: str, row, col, value, sheet_name: str = "Sheet1"):
    if col == -1:
        print('no key')
        return
    wb = load_workbook(path)
    wb[sheet_name].cell(row + 1, col + 1).value = value
    wb.save(path)

def key_index(data_frame, key_name):
    i = 0
    for key in data_frame.keys():
        if key == key_name:
            return i
        i += 1
    print("key wasn't found: " + key_name)
    return -1
